# This program will take a Commonwealth Honors Program student list with student IDs in the first column and
# add a preceding zero if it is missing and delete empty rows.
import openpyxl


def opener():
    global new_file
    global wb_obj
    global sheet_obj
    try:
        file_loc = input("Please enter the name of the Excel file, with the .xlsx at the end: ")
        new_file = input("Please enter the name of the converted file, with the .xlsx at the end: ")
        wb_obj = openpyxl.load_workbook(file_loc)
        sheet_obj = wb_obj.active

    except:
        print("That filename is invalid.")
        opener()


# This function will add a zero to the student ID numbers in the first column of an Excel spreadsheet.
def preceding_zero():
    row = sheet_obj.max_row
    for i in range(1, row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        first_digit = str(cell_obj.value)
        if first_digit != "None":
            if first_digit[0] != "0":
                if first_digit.isnumeric():
                    cell_obj.value = "0" + str(cell_obj.value)
                    print(cell_obj.value)
    wb_obj.save(new_file)


# based on code from https://stackoverflow.com/questions/52821618/openpyxl-how-to-delete-rows-from-an-excel-file-based-on-some-condition
def delete_empty_rows():
    i = 1
    while i <= sheet_obj.max_row:
        if sheet_obj.cell(row=i, column=1).value == None:
            sheet_obj.delete_rows(i, 1)
            # Note the absence of incremental.  Because we deleted a row, we want to stay on the same row because new data will show in the next iteration.
        else:
            i += 1
    wb_obj.save(new_file)


try:
    file_loc = input("""Put the file you want to convert in this folder: C:\\Users\\snowka\\PycharmProjects\\Excel_Tools \n Please enter the name of the Excel file, with the .xlsx at the end: """)
    new_file = input("Please enter the name of the converted file, with the .xlsx at the end: ")
    wb_obj = openpyxl.load_workbook(file_loc)
    sheet_obj = wb_obj.active

except:
    print("That filename is invalid.")
    opener()

preceding_zero()
remove_rows = input("Remove blank rows from the spreadsheet? Answer Y or N.")
if "Y" in remove_rows:
    delete_empty_rows()
elif "y" in remove_rows:
    delete_empty_rows()
elif "N" in remove_rows:
    pass
elif "n" in remove_rows:
    pass
else:
    print("That was not a valid response. Program closing.")
    pass
